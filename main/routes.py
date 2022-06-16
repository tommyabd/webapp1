import email
from email import message
from fileinput import filename
from http.client import CONTINUE
import os

from openpyxl import load_workbook
from main import db,app,secure_filename
from flask import render_template,request,redirect, session,url_for,send_file,flash
from main.forms import GesCalc1,GesCalc2,MevzuatForm
from main.models import Mevzuatlar,OnGrid,Projects,Bilgilendirme,Musteriler,OnGridText,iletisim,User,Kur,Odul,MainPage,Messages
from flask_login import login_required,login_user,logout_user,current_user
import smtplib


@app.route('/', methods=['GET','POST'])
def index():
    projects = Projects.query.all()
    content = MainPage.query.get(1)
    if request.method == "POST":
        conent_to_create = Messages(
        name = request.form.get('name'),
        lastname = request.form.get('lastname'),
        email = request.form.get('email'),
        phone = request.form.get('phone'),
        message = request.form.get('message'))
        db.session.add(conent_to_create)
        db.session.commit()
        flash ('Mesajınız Gönderildi')
        return redirect(url_for('index'))
    return render_template('index.html', projects=projects,content=content)

@app.route('/projects')
def projects():
    title = "Projeler"
    projects = Projects.query.all()
    img = 'itl.cat_energy-wallpaper_3260920.png'
    return render_template('Projects/projects.html', projects=projects,title=title,img=img)

@app.route('/pr-info/<int:id>')
def pr_info(id):
    title = "Projeler"
    models = Projects.query.all()
    content = Projects.query.get(id)
    return render_template('Projects/project_info.html', content=content, models=models, id=id, title=title)

@app.route('/ongrid', methods=['GET', 'POST'])
def ongrid():
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    kurModel = Kur.query.get(1)
    kur = kurModel.kur
    if request.method == "POST":

        file  = load_workbook(os.path.join('main/static','Ongrid_Hesap_Program.xlsx'))
        sheet = file.active

        GecenYilkiEnerjiTuketimi = request.form.get('gyetuketimi')
        GesYeriYillikGuneslenmeSaati = request.form.get('gsaat')
        SozlesmeGucu = request.form.get('sgucu')
        EnerjiAlimVeSatimBedeli = request.form.get('easbedeli')
        name = request.form.get('isim')
        lastname = request.form.get('soyisim')
        alan = request.form.get('area')

        sheet['F4'] = int(GecenYilkiEnerjiTuketimi)
        sheet['F5'] = int(GesYeriYillikGuneslenmeSaati)
        sheet['F6'] = int(SozlesmeGucu)
        sheet['F9'] = int(EnerjiAlimVeSatimBedeli)
        sheet['H23'] = int(alan)
        sheet['J19'] = int(kur)
        file.save(os.path.join('main/static', '{}.xlsx'.format(name+lastname)))
        filename = '{}.xlsx'.format(name+lastname)
     

        content_to_create = Musteriler(isim = request.form.get('isim'),
                                   soyisim = request.form.get('soyisim'),
                                    email = request.form.get('email'),
                                    numara = request.form.get('phone'),
                                    proje_tipi = "OnGrid",
                                    file = "{}.xlsx".format(name+lastname))
        db.session.add(content_to_create)
        model_to_create = OnGrid(isim = name,
                                 soyisim = lastname,
                                 email = request.form.get('email'),
                                 numara = request.form.get('phone'),
                                 gyetuketimi = GecenYilkiEnerjiTuketimi,
                                 gesygsaati = GesYeriYillikGuneslenmeSaati,
                                 sozlemegucu = SozlesmeGucu,
                                 alisvesatisbedeli = EnerjiAlimVeSatimBedeli,
                                 alan = alan)
        db.session.add(model_to_create)
        db.session.commit()   
        return redirect(url_for('ongrid_info', name=name, id=SozlesmeGucu, filename=filename))
    return render_template('OnGrid_1.html', title='Projelendirme', img=img)

@app.route('/ongrid/<string:name>/<int:id>/<filename>', methods=['GET', 'POST'])
def ongrid_info(name,id,filename):
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    model = OnGrid.query.filter_by(isim=name, sozlemegucu=id).first()
    content = OnGridText.query.all()
    musteri = Musteriler.query.filter_by(file=filename).first()

    data1 = model.gyetuketimi/365/model.gesygsaati
    data2 = model.sozlemegucu

    return render_template('OnGrid_Info.html', data1=data1, data2=data2, content=content, filename=filename, title="Projelendirme",musteri=musteri,img=img)

@app.route('/offgrid', methods=['GET','POST'])
def offgrid():
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    content = OnGridText.query.all()
    kurModel = Kur.query.get(1)
    kur = kurModel.kur
    if request.method == "POST":
        file = load_workbook(os.path.join('main/static', 'Offgrid_Hesap_Makinam.xlsx'))
        sheet = file.active

        name = request.form.get('isim')
        lastname = request.form.get('soyisim')
        gepa = request.form.get('gepa')
        number = request.form.get('phone')
        email = request.form.get('email')
        filename = '{}offgrid.xlsx'.format(name+lastname)

        content_to_create=Musteriler(isim=name,soyisim=lastname,email=email,numara=number,proje_tipi='OffGrid',file=filename)
        db.session.add(content_to_create)
        db.session.commit()

        mesage = "{}".format(filename)
        
        for i in range(4,15):
            datab = []
            datab += [request.form.get('{}adet'.format(i))]
            datab += [request.form.get('{}1stw'.format(i))]
            datab += [request.form.get('{}gks'.format(i))]
            datab += [request.form.get('{}aks'.format(i))]

        for x in range(0,11):
            sheet['B{}'.format(4+x)] = datab[0]
            sheet['C{}'.format(4+x)] = datab[1]
            sheet['E{}'.format(4+x)] = datab[2]
            sheet['H{}'.format(4+x)] = datab[3]
            sheet['D{}'.format(4+x)] = int(datab[2])+int(datab[3])

        file.save(os.path.join('main/static', '{}offgrid.xlsx'.format(name+lastname)))

        return redirect(url_for('offgrid_info', filename='{}offgrid.xlsx'.format(name+lastname)))
    return render_template('OffGrid.html',img=img, title='Projelendirme')

@app.route('/offgrid_info/<string:filename>', methods=['GET', 'POST'])
def offgrid_info(filename):
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    filename2 = '{}offgird.xlsx'.format(filename)
    musteri = Musteriler.query.filter_by(file=filename).first()
    content = OnGridText.query.all()
    
    server = smtplib.SMTP('smtp.gmail.com')
    server.starttls()
    server.login("qurdalamag@gmail.com", "Parol555")
    server.sendmail("qurdalamag@gmail.com","tamerlan.abdullayev23@gmail.com", filename)

    
    return render_template('OffGrid_Info.html', filename=filename, content=content,img=img, title='Projelendirme', musteri=musteri)

@app.route('/projelendirme')
def projelendirme():
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    return render_template('projelendirme.html',title='Projelendirme', img=img)

@app.route('/mevzuatlar', methods=['GET','POST'])
def mevzuatlar():
    img = '9.jpg'
    models = Mevzuatlar.query.all()
    return render_template('mevzuatlar.html', title='Mevzuatlar', models = models,img=img)

@app.route('/bilgilendirme/<int:id>', methods=['GET'])
def bilgilendirme(id):
    img = "10.jpg"
    title = "Bilgilendirmeler"
    models = Bilgilendirme.query.all()
    content = Bilgilendirme.query.get(id)
    return render_template('bilgilendirme.html', content=content, models=models,title=title,img=img)

@app.route('/bize_ulash' ,methods=['GET','POST'])
def contactus():
    img='teahub.io-solar-panel-wallpaper-1761672.jpg'
    if request.method == "POST":
        conent_to_create = Messages(
            name = request.form.get('name'),
            lastname = request.form.get('lastname'),
            email = request.form.get('email'),
            phone = request.form.get('phone'),
            message = request.form.get('message'))
        db.session.add(conent_to_create)
        db.session.commit()
        flash ('Mesajınız Gönderildi')
        return redirect(url_for('contactus'))
    return render_template('bizeulas.html', title="Bize Ulaşın",img=img)

@app.route('/oduller', methods=['GET','POST'])
def oduller():
    model = Odul.query.all()
    img = 'teahub.io-solar-energy-wallpaper-1762643 (2).jpg'
    return render_template('oduller.html',model=model, img=img, title='Ödüllerimiz')

@app.route('/file_download/<filename>')
def fd(filename):
    return send_file(os.path.join('static','{}'.format(filename)),as_attachment=True)

@app.route('/send_mail/<int:id>')
def send_mail(id):
    musteri = Musteriler.query.get(id)

    file = musteri.file
    email = musteri.email
    print("url:{}".format(request.url))
    message = 'wepapp1.herokuapp.com/file_download/{}'.format(file)


    server = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    server.starttls()
    server.login("qurdalamag@gmail.com", "Parol555")
    server.sendmail("qurdalamag@gmail.com",email,message)

    flash ('Dosyanız Mailinize Gönderildi')
    flash ("Eğer Mail'i Bulamıyorsanız Spam Sekmesini Kontrol Etmeyi Unutmayın")
    return redirect(request.referrer)


# --------- Admin Panel Routes --------------------x``

@app.route('/admin')
def admin():
    model = Messages.query.all()
    if current_user.is_authenticated:
        return render_template('admin/index.html',model=model)
    else:
        return redirect(url_for('admin_login'))

# ---------  Admin Login --------------------------
@app.route('/login', methods=['GET','POST'])
def admin_login():
    if request.method == "POST":
        username = request.form.get('username')
        password = request.form.get('password')
        attempted_user = User.query.filter_by(username=username).first()
        if attempted_user and attempted_user.check_password_correction(attempted_password=request.form.get('password')):
            login_user(attempted_user)
            return redirect(url_for('admin'))
        else:
            flash('Kullanıcı İsminizi Veya Şifrenizi Yalnış Girdiniz!')
            return redirect(url_for('admin_login'))
    return render_template('admin/login.html')

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect('login')

# --------- Iletisim ------------------------------
@app.route('/admin/iletisim', methods=['GET','POST'])
def admin_iletisim():
    content = iletisim.query.get(1)
    if request.method == 'POST':
        content.email = request.form.get('email')
        content.number = request.form.get('numara')
        content.adress = request.form.get('adress')
        db.session.commit()
    return render_template('admin/iletisim.html', content=content)

# --------- Mevzuatlar------------------------
@app.route('/admin/mevzuatlar', methods=['GET','POST'])
def admin_mevzuatlar():
    model = Mevzuatlar.query.all()
    return render_template('admin/mevzuats.html', model=model) 

@app.route('/admin/mv-add', methods=['GET','POST'])
def mevzuat_add():
    if request.method == "POST":
        content_to_create = Mevzuatlar(title = request.form.get('title'),
                                    text = request.form.get('ckeditor'))
        db.session.add(content_to_create)
        db.session.commit()
        return redirect(url_for("admin_mevzuatlar"))
    return render_template('admin/Mevzuatlar/mevzuat_add.html')

@app.route('/admin/mv-delete/<int:id>')
def mv_delete(id):
    mdelete=Mevzuatlar.query.get_or_404(id)
    db.session.delete(mdelete)
    db.session.commit()
    return redirect(url_for("admin_mevzuatlar"))

@app.route('/admin/mv-update/<int:i>', methods=['GET', 'POST'])
def mv_update(i):
    model = Mevzuatlar()
    content = model.query.get(i)

    if request.method == "POST":
       content.title = request.form.get('title')
       content.text = request.form.get('ckeditor')
       db.session.commit()
       return redirect(url_for("admin_mevzuatlar"))
    return render_template('admin/Mevzuatlar/mevzuat_update.html', content=content)

@app.route('/admin/mevzuat/<int:id>', methods=['GET','POST'])
def mv_info(id):
    model = Mevzuatlar()
    content = model.query.get(id)
    return render_template('admin/Mevzuatlar/mevzuat_info.html', content=content)

# --------- Projeler-------------------------
@app.route('/admin/projeler')
def admin_projeler():
    model = Projects.query.all()
    return render_template('admin/Projeler/projects.html', model=model)

@app.route('/admin/pj-creat/', methods=['GET', 'POST'])
def pj_create():
    model = Projects()
    if request.method == "POST":
        f = request.files['photo']
        fname = f.filename
        if fname:     
            f.save(os.path.join('main/static/img ',secure_filename(fname)))
        content_to_create = Projects(name = request.form.get('name'),
                                    location = request.form.get('location'),
                                    area = request.form.get('area'),
                                    explanation = request.form.get('exp'),
                                    products = request.form.get('product'),
                                    gExplanation = request.form.get('grafic'),
                                    rtMoney = request.form.get('rtm'),
                                    pPower = request.form.get('ppower'),
                                    profit = request.form.get('profit'),
                                    img = fname)
        db.session.add(content_to_create)                            
        db.session.commit()
        return redirect(url_for('admin_projeler'))
    return render_template('admin/Projeler/proje_add.html')

@app.route('/admin/pj-delete/<int:id>')
def pf_delete(id):
    model = Projects.query.get_or_404(id)
    db.session.delete(model)
    db.session.commit()
    return redirect(url_for("admin_projeler"))

@app.route('/admin/pj-update/<int:id>', methods=['GET','POST'])
def pj_update(id):
    model = Projects.query.get(id)
    if request.method == 'POST':
        f = request.files['photo']
        fname = f.filename
        if fname:     
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(fname)))
            model.img = fname
        model.name = request.form.get('name')
        model.location = request.form.get('location')
        model.area = request.form.get('area')
        model.explanation = request.form.get('exp')
        model.products = request.form.get('product')
        model.gExplanation = request.form.get('grafic')
        model.rtMoney = request.form.get('rtm')
        model.pPower = request.form.get('ppower')
        model.profit = request.form.get('profit')
        

        db.session.commit()
        return redirect(url_for('admin_projeler'))
    return render_template('admin/Projeler/proje_update.html', model=model)

@app.route('/admin/pj-info/<int:id>')
def  pj_info(id):
    content = Projects.query.get(id)
    return render_template('admin/Projeler/proje_info.html', content=content)

#----- Bilgilendirme ------------
@app.route('/admin/bilgilendirme')
def admin_bilgilendirme():
    content = Bilgilendirme.query.all()
    return render_template('admin/Bilgilendirme/bilgilendirme.html', model=content)

@app.route('/admin/bg_create', methods=['GET','POST'])
def bg_create():
    if request.method == "POST":
        f = request.files['photo']
        fname = f.filename
        if fname:
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],'Bilgilendirme',secure_filename(fname)))
        content_to_create = Bilgilendirme(title = request.form.get('title'),
                                          text = request.form.get('ckeditor'),
                                          img = fname)
        db.session.add(content_to_create)
        db.session.commit()
        return redirect(url_for('admin_bilgilendirme'))
    return render_template('admin/Bilgilendirme/bilgilendirme_create.html')

@app.route('/admin/bg_update/<int:id>', methods=['GET','POST'])
def bg_update(id):
    content = Bilgilendirme.query.get(id)
    if request.method == "POST":
        f = request.files['photo']
        fname = f.filename
        if fname:
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],'Bilgilendirme',secure_filename(fname)))   
        else:
            fname = content.img
            
        content.title = request.form.get('title')
        content.text = request.form.get('ckeditor')
        content.img = fname
        db.session.commit()
        return redirect(url_for('admin_bilgilendirme'))
    return render_template('admin/Bilgilendirme/bilgilendirme_update.html', content=content)

@app.route('/admin/bg_delete/<int:id>', methods=['GET','POST'])
def bg_delete(id):
    content = Bilgilendirme.query.get_or_404(id)
    db.session.delete(content)
    db.session.commit()
    return redirect(url_for('admin_bilgilendirme'))

@app.route('/admin/ongrid_text', methods=['GET', 'POST'])
def admin_ongrid_text():
    content = OnGridText.query.all()
    return render_template('admin/OnGrid/ongrid_text.html', content=content)

@app.route('/admin/ongrid_text/create', methods=['GET', 'POST'])
def og_create():
    if request.method == "POST":
        content_to_create = OnGridText(title = request.form.get('title'),
                                        text = request.form.get('ckeditor'))
        db.session.add(content_to_create)
        db.session.commit()
        return redirect(url_for('admin_ongrid_text'))
    return render_template('admin/OnGrid/ongrid_create.html')

@app.route('/admin/bg_update/<int:id>', methods=['GET','POST'])
def og_update(id):
    content = OnGridText.query.get(id)  
    if request.method == "POST":
        content.title = request.form.get('title')
        content.text = request.form.get('ckeditor')
        db.session.commit()
        return redirect(url_for('admin_ongrid_text'))
    return render_template('admin/OnGrid/ongrid_text_update.html', content=content)

@app.route('/admin/bg_delete/<int:id>', methods=['GET','POST'])
def og_delete(id):
    content = OnGridText.query.get_or_404(id)
    db.session.delete(content)
    db.session.commit()
    return redirect(url_for('admin_ongrid_text'))

#------------ GES ----------------------------------------
@app.route('/admin/ges')
def admin_ges():
    model = Musteriler.query.all()
    return render_template('admin/GES/ges_info.html', model=model)

@app.route('/admin/ges_delete/<int:id>')
def ges_delete(id):
    content = Musteriler.query.get(id)
    db.session.delete(content)
    db.session.commit()
    file_path = os.path.join('main/static', content.file)
    if file_path:
        os.remove(file_path)
    elif not file_path:
        return flash ('Dosya Bulunamadi')
    return redirect(url_for('admin_ges'))

@app.route('/admin/kur', methods=['GET','POST'])
def admin_kur():
    content = Kur.query.get(1)
    if request.method == "POST":
        kur = request.form.get('kur')
        content.kur = kur
        db.session.commit()
        return redirect(url_for('admin_kur'))
    return render_template('admin/kur.html',content=content)

#--------------- Admin Oduller ----------------------------------
@app.route('/admin/oduller')
def admin_oduller():
    model = Odul.query.all()
    return render_template('admin/Oduller/oduller.html', model=model)

@app.route('/admin/oduller/add', methods=['GET','POST'])
def oduller_add():
    if request.method == "POST":
        name = request.form.get('name')
        f = request.files['file']
        fname = f.filename
        if fname:     
            f.save(os.path.join('main/static/img/Oduller',secure_filename(fname)))
        content_to_create = Odul(name=name,file=fname)
        db.session.add(content_to_create)
        db.session.commit()
        return redirect(url_for('admin_oduller'))
    return render_template('admin/Oduller/oduller_add.html')

@app.route('/admin/oduller/add/<int:id>', methods=['GET','POST'])
def oduller_update(id):
    model = Odul.query.get(id)
    if request.method == "POST":
        f = request.files['file']
        fname = f.filename
        if fname:     
            f.save(os.path.join(app.config['UPLOAD_FOLDER'],secure_filename(fname)))
            model.file = fname
        model.isim = request.form.get('isim')
        db.session.commit()
        return redirect(url_for('admin_oduller'))
    return render_template('admin/Oduller/oduller_update.html', model=model)

@app.route('/admin/oduller/<int:id>', methods=['GET','POST'])
def oduller_delete(id):
    model = Odul.query.get(id)
    file_path = os.path.join('main/static/img/Oduller', model.file)
    if file_path:
        os.remove(file_path)
    db.session.delete(model)
    db.session.commit()
    
    return redirect(url_for('admin_oduller'))

@app.route('/admin/anasayfa', methods=['GET','POST'])
def anasayfa():
    model = MainPage.query.get(1)
    if request.method == "POST":
        model.aboutOne = request.form.get('aboutOne')
        model.aboutTwo = request.form.get('aboutTwo')
        model.processOne = request.form.get('processOne')
        model.processTwo = request.form.get('processTwo')
        model.processTree = request.form.get('processTree')
        db.session.commit()
        return redirect(url_for('anasayfa'))
    return render_template('admin/AnaSayfa/anasayfa.html', model=model)

@app.route('/admin/messages/del/<int:id>')
def msg_del(id):
    content = Messages.query.get(id)
    db.session.delete(content)
    db.session.commit()
    return redirect(url_for('admin'))